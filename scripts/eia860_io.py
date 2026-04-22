"""
Shared helpers for EIA Form 860 annual (xlsx) files.

Used by process-eia860.py and export-eia860-tables.py.
"""

from __future__ import annotations

import json
import os
import subprocess
import urllib.request
from datetime import date, datetime, time
from decimal import Decimal
from typing import Any

DEFAULT_EIA860_URL = (
    "https://www.eia.gov/electricity/data/eia860/xls/eia8602024.zip"
)


def ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("Installing openpyxl...")
        subprocess.check_call(["pip3", "install", "openpyxl", "-q"])
        import openpyxl  # noqa: F401


def normalize_cell(val: Any) -> Any:
    """Make values JSON-serializable (EIA xlsx uses dates, decimals, etc.)."""
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        return val
    if isinstance(val, Decimal):
        return float(val)
    if isinstance(val, datetime):
        return val.isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, time):
        return val.isoformat()
    return str(val)


def _rows_to_dicts(rows: list[tuple[Any, ...]], skip_rows: int) -> list[dict[str, Any]]:
    if len(rows) <= skip_rows:
        return []
    header_row = skip_rows
    headers = [
        str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[header_row])
    ]
    out: list[dict[str, Any]] = []
    for row in rows[header_row + 1 :]:
        d: dict[str, Any] = {}
        for i, val in enumerate(row):
            if i < len(headers):
                d[headers[i]] = normalize_cell(val)
        out.append(d)
    return out


def read_xlsx_sheet(filepath: str, sheet_name: str | None = None, skip_rows: int = 1):
    """Read one sheet; default is the workbook's active (first) sheet."""
    ensure_openpyxl()
    import openpyxl

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    return _rows_to_dicts(rows, skip_rows)


def read_xlsx_all_sheets(filepath: str, skip_rows: int = 1) -> dict[str, list[dict[str, Any]]]:
    """Read every sheet in a workbook; keys are sheet names."""
    ensure_openpyxl()
    import openpyxl

    out: dict[str, list[dict[str, Any]]] = {}
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            rows = list(ws.iter_rows(values_only=True))
            out[name] = _rows_to_dicts(rows, skip_rows)
    finally:
        wb.close()
    return out


def download_eia860_zip(
    url: str = DEFAULT_EIA860_URL,
) -> tuple[str, str]:
    """Download EIA 860 zip to a temp directory. Returns (tmp_dir, zip_path)."""
    import tempfile

    print("Downloading EIA Form 860 data...")
    tmp_dir = tempfile.mkdtemp()
    zip_path = os.path.join(tmp_dir, "eia860.zip")

    req = urllib.request.Request(
        url,
        headers={"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"},
    )
    with urllib.request.urlopen(req) as response:
        with open(zip_path, "wb") as f:
            f.write(response.read())

    file_size = os.path.getsize(zip_path)
    print(f"Downloaded {file_size / 1024 / 1024:.1f} MB to {zip_path}")
    return tmp_dir, zip_path


def write_json_compact(path: str, obj: Any) -> int:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, separators=(",", ":"), ensure_ascii=False)
    return os.path.getsize(path)


def safe_file_part(name: str) -> str:
    """Safe fragment for filenames derived from sheet titles."""
    return "".join(c if c.isalnum() or c in "-_" else "_" for c in str(name))
